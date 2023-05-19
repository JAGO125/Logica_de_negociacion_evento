from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import  os
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame
from operator import itemgetter
from openpyxl.worksheet.filters import SortCondition
from openpyxl.styles import Font
from openpyxl.formula import Tokenizer


# FUNCION PARA ELIMINAR ARCHIVO Convertido DESPUES DE CREADO
def eliminar_archivo():
    
    archivo_excel = './static/REPS/REPS_Clinica_Fatima_convertido.xlsx'

    if os.path.exists(archivo_excel):
        os.remove(archivo_excel)
        mensaje = f"El archivo {archivo_excel} ha sido eliminado exitosamente."
    else:
        mensaje = f"El archivo {archivo_excel} no existe."

    print (mensaje)

# FUNCION PARA CREAR Y ABRIR EL ARCHIVO REPS EN EXCEL

def abrir_REPS(request): #ABRIR EXCEL
    eliminar_archivo()   #SE LLAMA LA DEFINICION DENTRO
    ruta_archivo = './static/REPS/REPS_Clinica_Fatima.xlsx'
    libro_excel = load_workbook(ruta_archivo)
    

        # Imprimir los nombres de las hojas de cálculo
    """ for nombre in nombres_hojas:
        print(nombre)"""
    
    # CAMBIAR TIPO DE DATO
    nombre_hoja = 'Servicios (3)' 
    hoja = libro_excel[nombre_hoja] 
    numero_columna = 23  # Por ejemplo, para la columna A
    letra_columna = get_column_letter(numero_columna)

    for fila in hoja.iter_rows(min_row=2, min_col=numero_columna, max_col=numero_columna):
        #celda = fila[0]
        #if celda.data_type == "s":
            #celda.data_type = "n"  # Convertir a tipo numérico

        for celda in fila:
                if celda.data_type == 's':
                    continue
                if celda.data_type == 'n':
                    celda.number_format = numbers.FORMAT_TEXT   #CAMBIAR TIPO DE DATO

    #INSERTAR NUEVA COLUMNA

    columna_ah_index = hoja['AH'][0].column  # Obtiene el índice de la columna AH
    #columna_ah_index = 34
    columna_ai_index = hoja['AI'][0].column  # Obtiene el índice de la columna AI
    #columna_ai_index = 35
    

    nueva_columna_index = columna_ah_index + 1  # Calcula el índice de la nueva columna
    
    hoja.insert_cols(nueva_columna_index)  # Inserta la nueva columna

    # Actualiza las posiciones de las columnas restantes
    for fila in hoja.iter_rows():
        for celda in fila[nueva_columna_index:]:
            celda.column = celda.column

    #AGREGAR TITULO A UNA COLUMNA

   # nueva_columna_letra = get_column_letter(nueva_columna_index)  # Obtiene la letra de la columna

    nombre_columna = "NIVEL"  # Nombre que deseo asignar a la columna

    celda_nombre_columna = hoja.cell(row=1, column=nueva_columna_index)
    celda_nombre_columna.value = nombre_columna

    # APLICAR FORMULA A UNA CELDA

    
    columna = hoja['AI']  # Columna donde se aplicará la fórmula

    for celda in columna[1:]:
        #=SI(AH6="SI";3;SI(AG6="SI";2;SI(AF6="SI";1;SI(BL6="ALTA";3;SI(BL6="MEDIANA";2;SI(BL6="BAJA";1;1))))))
        #=SI(AH{}="SI";3;SI(AG{}="SI";2;SI(AF{}="SI";1;SI(BL{}="ALTA";3;SI(BL{}="MEDIANA";2;SI(BL{}="BAJA";1;1))))))
        #celda.value = '=IF(AH{}="SI",3,IF(AG{}="SI",2,1))'.format(celda.row, celda.row)
        celda.value = '=IF(AH{}="SI",3,IF(AG{}="SI",2,IF(AF{}="SI",1,IF(BL{}="ALTA",3,IF(BL{}="MEDIANA",2,IF(BL{}="BAJA",1,1))))))'.format(celda.row, celda.row, celda.row, celda.row, celda.row, celda.row)

    # Actualiza las referencias de las celdas en las fórmulas
    #for celda in hoja.iter_cells():
     #   if celda.data_type == 'f':
      #      celda.value = celda.value.replace(f"{get_column_letter(columna_ai_index)}", f"{get_column_letter(nueva_columna_index)}")                

    # Obtener los valores de la columna W

    # Obtener los valores de la columna W junto con sus filas correspondientes
    # Obtener los valores de la columna W junto con sus filas correspondientes


    rango_w = hoja["W2:W" + str(hoja.max_row)]  # Rango de celdas de la columna W, desde la fila 2 hasta la última fila
    valores_w = [(celda.value, celda.row) for fila in rango_w for celda in fila]
    

    # Ordenar los valores de la columna W de mayor a menor
    valores_w.sort(reverse=True, key=itemgetter(0))
    print(valores_w)

    # Actualizar los valores en la columna W de la hoja
    

        

    ruta_guardado = './static/REPS/REPS_Clinica_Fatima_convertido.xlsx'
    libro_excel.save(ruta_guardado)
    

    
         

    return HttpResponse("El archivo se cargó correctamente")


def eliminar_maestra():
    
    archivo_excel = './static/MAESTRA/230413 - ListadoTecnologiasEmssanar_Convertido.xlsx'

    if os.path.exists(archivo_excel):
        os.remove(archivo_excel)
        mensaje = f"El archivo {archivo_excel} ha sido eliminado exitosamente."
    else:
        mensaje = f"El archivo {archivo_excel} no existe."

    print (mensaje)


def abrir_MAESTRA(request):
    eliminar_maestra()
    ruta_archivo = './static/MAESTRA/230413 - ListadoTecnologiasEmssanar.xlsx'
    libro_excel = load_workbook(ruta_archivo)

    hoja_tecnologia = libro_excel["TECNOLOGIAS ( CUPS )"]

    # Eliminar la hoja "tecnologia"
    libro_excel.remove(hoja_tecnologia)

    hoja_tecnologia = libro_excel["Cod_Derogados & HomologacionT"]

    # Eliminar la hoja "tecnologia"
    libro_excel.remove(hoja_tecnologia)

    #INSERTAR NUEVA COLUMNA
  

    # Obtener la hoja de cálculo
    hoja = libro_excel["SERVICIOS_HABILITADOS"]

    # Obtener la letra de la última columna existente
    ultima_columna = hoja.max_column
    letra_ultima_columna = get_column_letter(ultima_columna)

    # Insertar las nuevas columnas al final
    hoja.insert_cols(ultima_columna + 1)
    hoja.insert_cols(ultima_columna + 2)
    hoja.insert_cols(ultima_columna + 3)
    hoja.insert_cols(ultima_columna + 4)

    # Escribir los encabezados de las nuevas columnas
    hoja.cell(row=1, column=ultima_columna + 1, value="SERV_HAB").font = Font(bold=True)
    hoja.cell(row=1, column=ultima_columna + 2, value="VALOR").font = Font(bold=True)
    hoja.cell(row=1, column=ultima_columna + 3, value="NIVEL_IPS").font = Font(bold=True)
    hoja.cell(row=1, column=ultima_columna + 4, value="VALIDAR_NIVEL").font = Font(bold=True)

    archivo_excel2 = "./static/REPS/REPS_Clinica_Fatima_convertido_final.xlsx"
    libro2 = load_workbook(archivo_excel2)

    # Obtener la hoja de cálculo del segundo archivo
    hoja2 = libro2["Servicios (3)"]

    # Obtener el primer dato de la columna G del primer archivo
    primer_dato_g = hoja["G2"].value

    # Realizar la fórmula BUSCARV en todos los datos de la columna W del segundo archivo
    letra_w = get_column_letter(hoja2["W"].column)
    ultima_fila = hoja2.max_row

    for fila in range(2, ultima_fila + 1):
        celda_w = hoja2[letra_w + str(fila)]
        formula = '=BUSCARV({}, {}2:{}{}, 1, FALSO)'.format(primer_dato_g, letra_w, letra_w, ultima_fila)
        celda_w.value = formula   

    ruta_guardado = './static/MAESTRA/230413 - ListadoTecnologiasEmssanar_Convertido.xlsx'
    libro_excel.save(ruta_guardado)

    return HttpResponse("La Maestra se cargó correctamente")





    
